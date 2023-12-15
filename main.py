from datetime import datetime, timedelta
from tkinter import Tk, filedialog, simpledialog, ttk
import openpyxl

def process_excel_file(file_path, output_file_path, specified_date):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    with open(output_file_path, 'w', encoding='utf-8') as output_file:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            col2_value = row[1]
            col1_value = row[0]
            col8_value = row[7]

            result_string = f"01{col2_value}21{col1_value}§17{specified_date.strftime('%y%m%d')}10{col8_value}"

            result_string = result_string.replace('"', '').replace('=', '')

            output_file.write(f"{result_string}\n")

    print("Обработка завершена. Результат сохранен в", output_file_path)

def save_to_txt(output_file_path, data):
    with open(output_file_path, 'w', encoding='utf-8') as output_file:
        output_file.write(data)

def save_to_excel(output_file_path, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    rows = data.split('\n')
    for row in rows:
        sheet.append([row])

    workbook.save(output_file_path)

def select_file():
    root = Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Excel Files", "*.xlsx;*.xls")])

    return file_path

def select_date():
    root = Tk()
    root.withdraw()

    year_month = simpledialog.askstring("Год и месяц", "Введите год и месяц (гг-мм):")

    try:
        specified_date = datetime.strptime(f"{year_month}-01", "%y-%m-%d")
    except ValueError:
        print("Неверный формат года и месяца. Используйте формат гг-мм.")
        return None

    specified_date = (specified_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)

    return specified_date

def get_output_file_info():
    root = Tk()
    root.withdraw()

    file_info = {}

    # Ввод названия файла
    file_info['name'] = simpledialog.askstring("Название файла", "Введите название файла для сохранения:")
    if not file_info['name']:
        print("Название файла не введено. Используется стандартное название.")
        file_info['name'] = "Штрихкоды_Киргизия"

    # Выбор формата файла
    file_info['format'] = ttk.Combobox(root, values=['txt', 'excel'])
    file_info['format'].set('txt')
    file_info['format'].pack(pady=10)
    file_info['format_label'] = ttk.Label(root, text="Выберите формат файла:")
    file_info['format_label'].pack(pady=5)

    root.mainloop()

    return file_info

if __name__ == "__main__":
    excel_file_path = select_file()
    if not excel_file_path:
        print("Отменено пользователем.")
    else:
        specified_date = select_date()

        if specified_date:
            output_file_info = get_output_file_info()
            output_file_path = f"{output_file_info['name']}.{output_file_info['format'].get()}"

            if output_file_info['format'].get() == 'txt':
                process_excel_file(excel_file_path, output_file_path, specified_date)
            elif output_file_info['format'].get() == 'excel':
                data = process_excel_file(excel_file_path, specified_date)
                save_to_excel(output_file_path, data)
